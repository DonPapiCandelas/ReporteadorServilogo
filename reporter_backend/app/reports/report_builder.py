# app/reports/report_builder.py
import io
import datetime
import decimal
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, PageBreak, Spacer, Image as RLImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas

# Importamos los esquemas que definimos
from .report_schemas import CurrencyGroup, AgingSummary

# --- Tus Helpers de Estilo Originales ---
THEME = {
    "primary": "2E86AB",
    "secondary": "A7C957",
    "head": "2E86AB",
    "bg_soft": "F8FAFC",
    "line": "D1D5DB",
    "ink": "1F2937",
    "total": "D1EFB5",
}

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def border_all():
    edge = Side(style="thin", color=THEME["line"])
    return Border(left=edge, right=edge, top=edge, bottom=edge)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

def insert_logo(ws, tl_cell, logo_path, max_w=170, max_h=42):
    # Por ahora, no implementamos el logo_path, pero dejamos la función
    pass
    # try:
    #     if not logo_path or not Path(logo_path).exists():
    #         return
    #     img = XLImage(logo_path)
    #     scale = min(max_w / img.width, max_h / img.height, 1.0)
    #     img.width = int(img.width * scale)
    #     img.height = int(img.height * scale)
    #     ws.add_image(img, tl_cell)
    # except Exception as e:
    #     logging.warning(f"No se pudo insertar el logo de Excel: {e}")

def _safe_excel_title(s: str) -> str:
    bad = '[]:*?/\\'
    for ch in bad:
        s = s.replace(ch, ' ')
    return (s[:31] or "Sheet").rstrip()

def fmt_date(v, out_fmt="%m/%d/%y"):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime(out_fmt)
    return str(v) if v is not None else ""

def _ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"

def fmt_date_ordinal(d: datetime.date) -> str:
    if isinstance(d, datetime.datetime):
        d = d.date()
    return f"{d.strftime('%B')} {_ordinal(d.day)}, {d.year}"

def _rl_logo(logo_path, max_w=30 * mm, max_h=12 * mm):
    # Por ahora, no implementamos el logo_path
    return None

def _pdf_header_footer(canv: rl_canvas.Canvas, doc):
    canv.saveState()
    canv.setFont("Helvetica", 7)
    w, h = landscape(A4)
    canv.drawString(8 * mm, 8 * mm, f"Generado: {datetime.date.today():%Y-%m-%d}")
    canv.drawRightString(w - 8 * mm, 8 * mm, f"Página {doc.page}")
    canv.restoreState()

# --- La Lógica de 'build_excel' (Adaptada) ---

def create_excel_report(
    data: Dict[str, CurrencyGroup], 
    logo_path: str, # Aún no lo usamos, pero lo pasamos
    filters: dict
) -> io.BytesIO:
    """
    Toma los datos procesados y construye un archivo Excel EN MEMORIA.
    Devuelve un objeto io.BytesIO.
    """

    as_of = filters['as_of']
    customer_name = filters['customer_name']
    is_single_customer = filters.get('customer_id') is not None
    
    all_entries = [entry for group in data.values() for entry in group.entries]

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # ===== 1. Hojas por moneda (Currency Sheets) - PRIMERAS =====
    for cur, cur_group in data.items():
        cur_rows = cur_group.entries

        ws2 = wb.create_sheet(_safe_excel_title(f"CURRENCY {cur}"))
        insert_logo(ws2, "A1", logo_path)
        
        merge_end = "J1" if is_single_customer else "K1"
        merge_end_row2 = "J2" if is_single_customer else "K2"
        
        ws2.merge_cells(f"C1:{merge_end}")
        a1 = ws2["C1"]
        a1.value = f"Receivables Aging — {cur}"
        a1.font = Font(bold=True, size=16, color=THEME["primary"])
        a1.alignment = Alignment(horizontal="center")
        
        ws2.merge_cells(f"C2:{merge_end_row2}")
        a2 = ws2["C2"]
        a2.value = f"As Of: {as_of:%m/%d/%Y} • Customer: {customer_name}"
        a2.font = Font(color=THEME["ink"])
        a2.alignment = Alignment(horizontal="center")

        if is_single_customer:
            headers2 = [
                "REFERENCE", "DOCUMENT", "NO.", "INVOICE DATE", 
                "TOTAL AMOUNT", "ARRIVAL DATE", "PAYMENTS", "BALANCE", 
                "DUE DATE", "DAYS SINCE ARRIVAL"
            ]
        else:
            headers2 = [
                "CUSTOMER", "REFERENCE", "DOCUMENT", "NO.", "INVOICE DATE", 
                "TOTAL AMOUNT", "ARRIVAL DATE", "PAYMENTS", "BALANCE", 
                "DUE DATE", "DAYS SINCE ARRIVAL"
            ]

        for i, h in enumerate(headers2, start=1):
            c = ws2.cell(4, i, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill = fill(THEME["head"])
            c.border = border_all()

        r0 = 5
        for idx, entry in enumerate(cur_rows, start=r0):
            if is_single_customer:
                out = [
                    entry.reference, entry.module, entry.folio, entry.invoice_date,
                    entry.total, entry.arrival_date, entry.paid, entry.balance,
                    entry.due_date, entry.days_since
                ]
            else:
                out = [
                    entry.customer_name, entry.reference, entry.module, entry.folio, entry.invoice_date,
                    entry.total, entry.arrival_date, entry.paid, entry.balance,
                    entry.due_date, entry.days_since
                ]
            
            for i, v in enumerate(out, start=1):
                if isinstance(v, (datetime.datetime, datetime.date)):
                    ws2.cell(idx, i).value = v
                elif isinstance(v, (int, float, decimal.Decimal)):
                    ws2.cell(idx, i).value = v
                else:
                    ws2.cell(idx, i).value = str(v) if v is not None else ""

        if cur_rows:
            last2 = r0 + len(cur_rows) - 1
            _apply_currency_sheet_formats_excel_custom(ws2, r0, last2, is_single_customer)
            
            if is_single_customer:
                widths = [30, 14, 8, 12, 16, 14, 16, 16, 14, 14]
            else:
                widths = [28, 30, 14, 8, 12, 16, 14, 16, 16, 14, 14]
            set_col_widths(ws2, widths)

    # ===== 2. Hojas Summary (Summary Sheets) - SEGUNDAS =====
    for cur, cur_group in data.items():
        ws3 = wb.create_sheet(_safe_excel_title(f"SUMMARY {cur}"))
        insert_logo(ws3, "A1", logo_path)
        ws3.merge_cells("C1:H1")
        s1 = ws3["C1"]
        s1.value = f"ACCOUNTS RECEIVABLE — SUMMARY ({cur})"
        s1.font = Font(bold=True, size=14, color=THEME["primary"])
        s1.alignment = Alignment(horizontal="center")
        ws3.merge_cells("C2:H2")
        s2 = ws3["C2"]
        s2.value = f"As Of: {as_of:%m/%d/%Y} • Customer: {customer_name}"
        s2.alignment = Alignment(horizontal="center")

        if is_single_customer:
            hdr = [
                "TOTAL BALANCE", "NOT YET DUE", "OVERDUE",
                "0-21", "22-30", "31-45", "45+ DAYS",
            ]
        else:
            hdr = [
                "CUSTOMER", "TOTAL BALANCE", "NOT YET DUE", "OVERDUE",
                "0-21", "22-30", "31-45", "45+ DAYS",
            ]

        for i, h in enumerate(hdr, start=1):
            c = ws3.cell(3, i, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.fill = fill(THEME["head"])
            c.border = border_all()

        r = 4
        for cust, agg in sorted(cur_group.aging_summary.items()):
            col_idx = 1
            if not is_single_customer:
                ws3.cell(r, col_idx).value = cust
                col_idx += 1
            
            vals = [
                agg.total_balance, agg.not_yet_due, agg.overdue,
                agg.bucket_0_21, agg.bucket_22_30, agg.bucket_31_45, agg.bucket_45_plus
            ]
            for v in vals:
                ws3.cell(r, col_idx).value = v
                col_idx += 1
            r += 1

        if r > 4:
            last3 = r - 1
            # Ajustamos el rango de columnas según si hay cliente o no
            start_col_fmt = 1 if is_single_customer else 2
            end_col_fmt = 7 if is_single_customer else 8
            
            _apply_summary_formats_excel(ws3, 4, last3, is_single_customer)
            
            if is_single_customer:
                set_col_widths(ws3, [18, 18, 18, 12, 12, 12, 12])
            else:
                set_col_widths(ws3, [30, 18, 18, 18, 12, 12, 12, 12])

    # ===== 3. Main Report - AL FINAL =====
    ws = wb.create_sheet("Main Report")
    
    for r in range(1, 6):
        for c in range(1, 15):
            ws.cell(r, c).fill = fill(THEME["bg_soft"])
    insert_logo(ws, "A1", logo_path)
    ws.merge_cells("C1:N2")
    t = ws["C1"]
    t.value = "ACCOUNTS RECEIVABLE AGING REPORT"
    t.font = Font(name="Segoe UI Semibold", size=18, color=THEME["primary"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws["C4"].value = "AS OF:"
    ws["C4"].font = Font(bold=True)
    ws["D4"].value = as_of
    ws["D4"].number_format = "mmmm d, yyyy"
    ws["G4"].value = "CUSTOMER:"
    ws["G4"].font = Font(bold=True)
    ws["H4"].value = customer_name
    for a in ("C4", "D4", "G4", "H4"):
        ws[a].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "CUSTOMER", "REFERENCE", "DOCUMENT", "INVOICE DATE", "NO.",
        "ARRIVAL DATE", "DUE DATE", "CURRENCY", "FX RATE", "TOTAL AMOUNT",
        "PAYMENTS", "BALANCE", "DAYS SINCE ARRIVAL",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(6, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill(THEME["head"])
        c.border = border_all()

    start = 7
    for r, entry in enumerate(all_entries, start=start):
        out = [
            entry.customer_name, entry.reference, entry.module, entry.invoice_date,
            entry.folio, entry.arrival_date, entry.due_date, entry.currency,
            entry.fx_rate, entry.total, entry.paid, entry.balance, entry.days_since,
        ]
        for i, v in enumerate(out, start=1):
            if isinstance(v, (datetime.datetime, datetime.date)):
                ws.cell(r, i).value = v
            elif isinstance(v, (int, float, decimal.Decimal)):
                ws.cell(r, i).value = v
            else:
                ws.cell(r, i).value = str(v) if v is not None else ""

    last = start + len(all_entries) - 1 if all_entries else start
    if all_entries:
        _apply_main_report_formats_excel(ws, start, last)

    set_col_widths(ws, [28, 30, 14, 12, 8, 14, 14, 10, 10, 16, 16, 16, 14])

    if len(wb.sheetnames) > 0:
        wb.active = 0

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook

# --- Helpers de formato de Excel ---

def _apply_main_report_formats_excel(ws, start, last):
    if start > last: return
    for row in range(start, last + 1):
        ws.cell(row, 4).number_format = "mm/dd/yyyy"
        ws.cell(row, 6).number_format = "mm/dd/yyyy"
        ws.cell(row, 7).number_format = "mm/dd/yyyy"
        ws.cell(row, 9).number_format = "0.0000"
        for col in (10, 11, 12):
            ws.cell(row, col).number_format = "$#,##0.00"
        ws.cell(row, 13).number_format = "0"
        ws.cell(row, 1).alignment = Alignment(horizontal="left")
        ws.cell(row, 2).alignment = Alignment(horizontal="left")
        for col in range(3, 14):
            ws.cell(row, col).alignment = Alignment(horizontal="center")
        for col in range(1, 14):
            ws.cell(row, col).border = border_all()
    tr = last + 1
    ws.merge_cells(f"A{tr}:I{tr}")
    ws.cell(tr, 1).value = "TOTALS:"
    ws.cell(tr, 1).alignment = Alignment(horizontal="right")
    for col in (10, 11, 12):
        ws.cell(tr, col).value = f"=SUM({chr(64 + col)}{start}:{chr(64 + col)}{last})"
        ws.cell(tr, col).number_format = "$#,##0.00"
        ws.cell(tr, col).font = Font(bold=True)
        ws.cell(tr, col).fill = fill(THEME["total"])
        ws.cell(tr, col).alignment = Alignment(horizontal="center")
        ws.cell(tr, col).border = border_all()
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:M{last}"
    overdue_fill = PatternFill("solid", fgColor="FDE68A")
    ws.conditional_formatting.add(
        f"M{start}:M{last}",
        CellIsRule(operator='greaterThan', formula=['45'], stopIfTrue=True, fill=overdue_fill)
    )

def _apply_currency_sheet_formats_excel_custom(ws2, r0, last2, is_single_customer):
    if r0 > last2: return
    
    if is_single_customer:
        date_cols = (4, 6, 9)
        money_cols = (5, 7, 8)
        int_cols = (10,)
        total_cols = (5, 7, 8)
        days_col = 10
        last_col = 10
        merge_range_end = 4
    else:
        date_cols = (5, 7, 10)
        money_cols = (6, 8, 9)
        int_cols = (11,)
        total_cols = (6, 8, 9)
        days_col = 11
        last_col = 11
        merge_range_end = 5

    for row in range(r0, last2 + 1):
        for col in date_cols:
            ws2.cell(row, col).number_format = "mm/dd/yyyy"
        for col in money_cols:
            ws2.cell(row, col).number_format = "$#,##0.00"
        for col in int_cols:
            ws2.cell(row, col).number_format = "0"
            
        ws2.cell(row, 1).alignment = Alignment(horizontal="left")
        if not is_single_customer:
            ws2.cell(row, 2).alignment = Alignment(horizontal="left")
        
        for col in range(1, last_col + 1):
            if col > (2 if not is_single_customer else 1):
                ws2.cell(row, col).alignment = Alignment(horizontal="center")
            ws2.cell(row, col).border = border_all()

    tr2 = last2 + 1
    merge_char = chr(64 + merge_range_end)
    ws2.merge_cells(f"A{tr2}:{merge_char}{tr2}")
    ws2.cell(tr2, 1).value = f"TOTAL ({ws2.title.split(' ', 1)[-1]}):"
    ws2.cell(tr2, 1).alignment = Alignment(horizontal="right")
    
    for col in total_cols:
        ws2.cell(tr2, col).value = f"=SUM({chr(64 + col)}{r0}:{chr(64 + col)}{last2})"
        ws2.cell(tr2, col).number_format = "$#,##0.00"
        ws2.cell(tr2, col).font = Font(bold=True)
        ws2.cell(tr2, col).fill = fill(THEME["total"])
        ws2.cell(tr2, col).alignment = Alignment(horizontal="center")
        ws2.cell(tr2, col).border = border_all()
        
    ws2.freeze_panes = "A5"
    last_col_char = chr(64 + last_col)
    ws2.auto_filter.ref = f"A4:{last_col_char}{last2}"
    
    overdue_fill = PatternFill("solid", fgColor="FDE68A")
    days_col_char = chr(64 + days_col)
    ws2.conditional_formatting.add(
        f"{days_col_char}{r0}:{days_col_char}{last2}",
        CellIsRule(operator='greaterThan', formula=['45'], stopIfTrue=True, fill=overdue_fill)
    )

def _apply_summary_formats_excel(ws3, start, last, is_single_customer=False):
    # Definir columnas según si hay cliente o no
    if is_single_customer:
        money_cols = range(1, 8) # 1 to 7
        total_cols = range(1, 8) # 1 to 7
        border_cols = range(1, 8)
    else:
        money_cols = range(2, 9) # 2 to 8
        total_cols = range(2, 9) # 2 to 8
        border_cols = range(1, 9)

    for row in range(start, last + 1):
        if not is_single_customer:
            ws3.cell(row, 1).alignment = Alignment(horizontal="left")
        
        for col in money_cols:
            ws3.cell(row, col).number_format = "$#,##0.00"
            ws3.cell(row, col).alignment = Alignment(horizontal="center")
        
        for col in border_cols:
            ws3.cell(row, col).border = border_all()

    ws3[f"A{last + 1}"].value = "TOTALS:"
    ws3[f"A{last + 1}"].font = Font(bold=True)
    
    # Si es single customer, TOTALS está en A (col 1), pero los valores también empiezan en col 1?
    # No, si es single customer, la col 1 es Total Balance.
    # Entonces "TOTALS:" debería ir en una fila antes o en una celda separada?
    # En el diseño original, TOTALS iba en A (Customer column).
    # Si quitamos Customer, la col 1 es Total Balance.
    # Entonces no podemos poner "TOTALS:" en A{last+1} si ahí va un número.
    # Pero wait, la fila de totales es una fila APARTE.
    # En la fila de totales, queremos sumar las columnas.
    # Si es single customer, col 1 es Total Balance.
    # Donde ponemos la etiqueta "TOTALS:"?
    # Quizás no ponemos etiqueta o la ponemos en un header superior?
    # O insertamos una fila antes?
    # En el código original:
    # ws3[f"A{last + 1}"].value = "TOTALS:"
    # Y luego itera col 2..8 para poner las sumas.
    # Si es single customer, col 1..7 son sumas.
    # Entonces "TOTALS:" sobrescribiría la suma de Total Balance si lo ponemos en A{last+1}.
    # Solución: Poner "TOTALS:" en una celda combinada o simplemente omitirlo en la misma celda, 
    # o moverlo a la izquierda si hubiera espacio (no hay).
    # O quizás ponerlo en la fila anterior?
    # Vamos a dejarlo sin etiqueta "TOTALS:" en la celda de datos, o ponerlo en el header?
    # Mejor: Si es single customer, NO ponemos etiqueta "TOTALS:" en la columna 1, porque ahí va un dato.
    # Simplemente ponemos los totales.
    
    if not is_single_customer:
        ws3[f"A{last + 1}"].value = "TOTALS:"
        ws3[f"A{last + 1}"].font = Font(bold=True)

    for col in total_cols:
        L = chr(64 + col)
        cell = ws3[f"{L}{last + 1}"]
        cell.value = f"=SUM({L}{start}:{L}{last})"
        cell.number_format = "$#,##0.00"
        cell.font = Font(bold=True)
        cell.fill = fill(THEME["total"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all()

# --- NUEVA FUNCIÓN DE PDF ---

def create_pdf_report(
    data: Dict[str, CurrencyGroup], 
    logo_path: str, # Aún no lo usamos
    filters: dict
) -> io.BytesIO:
    """
    Toma los datos procesados y construye un archivo PDF EN MEMORIA.
    Devuelve un objeto io.BytesIO.
    """

    as_of = filters['as_of']
    customer_name = filters['customer_name']
    # Detectar si es un solo cliente
    is_single_customer = filters.get('customer_id') is not None

    # --- ¡CAMBIO CLAVE! Guardar en memoria ---
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=8 * mm, rightMargin=8 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"], fontSize=12, textColor=colors.HexColor("#2E86AB"), spaceAfter=2 * mm
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#6B7280"), spaceAfter=4 * mm
    )

    story = []
    head_color = colors.HexColor("#2E86AB")
    line_color = colors.HexColor("#D1D5DB")
    alt_color = colors.HexColor("#F8F9FA")
    total_color = colors.HexColor("#D1EFB5")

    for idx, (cur, cur_group) in enumerate(data.items()):
        if idx > 0:
            story.append(PageBreak())

        lg = _rl_logo(logo_path)
        if lg:
            story.append(lg)
            story.append(Spacer(2 * mm, 0))

        story.append(Paragraph(f"CURRENCY — {cur}", title_style))
        story.append(
            Paragraph(
                f"As Of: <b>{as_of:%m/%d/%y}</b> &nbsp;|&nbsp; Customer: <b>{customer_name}</b> &nbsp;|&nbsp; Records: <b>{len(cur_group.entries)}</b>",
                meta_style,
            )
        )

        # --- Definición de Columnas (Currency) ---
        if is_single_customer:
            # Sin columna Customer
            # Excel Order: REFERENCE, DOCUMENT, NO., INVOICE DATE, TOTAL AMOUNT, ARRIVAL DATE, PAYMENTS, BALANCE, DUE DATE, DAYS SINCE ARRIVAL
            headers = [
                "REFERENCE", "DOC", "NO.", "INV\nDATE", "TOTAL", "ARR\nDATE",
                "PAYMT", "BALANCE", "DUE\nDATE", "DAYS",
            ]
            # Ajustamos anchos
            widths = [80, 14, 12, 14, 25, 14, 22, 26, 14, 18] 
        else:
            # Excel Order: CUSTOMER, REFERENCE, DOCUMENT, NO., INVOICE DATE, TOTAL AMOUNT, ARRIVAL DATE, PAYMENTS, BALANCE, DUE DATE, DAYS SINCE ARRIVAL
            headers = [
                "CUSTOMER", "REFERENCE", "DOC", "NO.", "INV\nDATE", "TOTAL", "ARR\nDATE",
                "PAYMT", "BALANCE", "DUE\nDATE", "DAYS",
            ]
            widths = [50, 60, 14, 12, 14, 25, 14, 22, 26, 14, 18]

        tbl_data = [headers]
        for entry in cur_group.entries:
            row_vals = []
            
            # 1. Customer (Solo si NO es single customer)
            if not is_single_customer:
                customer_display = entry.customer_name or ""
                if len(customer_display) > 25:
                    customer_display = customer_display[:25] + "\n" + customer_display[25:45]
                row_vals.append(customer_display)

            # 2. Reference
            reference_display = entry.reference or ""
            if len(reference_display) > 30:
                reference_display = reference_display[:30] + "\n" + reference_display[30:55]
            row_vals.append(reference_display)

            # 3. Doc Type
            doc_abbr = (
                "Inv" if entry.module == "Invoice"
                else "CrNote" if entry.module == "Credit Note"
                else "SO" if entry.module == "Sales Order"
                else "Pmt" if entry.module == "Customer Payment"
                else (entry.module or "")[:5]
            )
            row_vals.append(doc_abbr)

            # 4. No. (Folio)
            row_vals.append(str(entry.folio or ""))

            # 5. Invoice Date
            row_vals.append(fmt_date(entry.invoice_date, "%m/%d/%y"))

            # 6. Total Amount
            row_vals.append(f"{entry.total:,.2f}")

            # 7. Arrival Date
            row_vals.append(fmt_date(entry.arrival_date, "%m/%d/%y"))

            # 8. Payments
            row_vals.append(f"{entry.paid:,.2f}")

            # 9. Balance
            row_vals.append(f"{entry.balance:,.2f}")

            # 10. Due Date
            row_vals.append(fmt_date(entry.due_date, "%m/%d/%y"))

            # 11. Days
            row_vals.append(str(entry.days_since))
            
            tbl_data.append(row_vals)

        t = Table(tbl_data, colWidths=[w * mm for w in widths], repeatRows=1, splitByRow=True)
        
        # Estilos dinámicos según columnas
        # Indices de columnas numéricas (Total, Pmt, Bal) dependen de si hay Customer o no
        # Si is_single_customer: 
        #   0: Ref, 1: Doc, 2: No, 3: InvDate, 4: Total, 5: ArrDate, 6: Pmt, 7: Bal, 8: DueDate, 9: Days
        #   Align Left: 0 (Ref)
        #   Align Center: 1, 2, 3, 5, 8, 9
        #   Align Right: 4, 6, 7
        
        # Si NO single customer:
        #   0: Cust, 1: Ref, 2: Doc, 3: No, 4: InvDate, 5: Total, 6: ArrDate, 7: Pmt, 8: Bal, 9: DueDate, 10: Days
        #   Align Left: 0, 1
        #   Align Center: 2, 3, 4, 6, 9, 10
        #   Align Right: 5, 7, 8
        
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), head_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 6),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 6),
            ("GRID", (0, 0), (-1, -1), 0.25, line_color),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, alt_color]),
            ("LINEBELOW", (0, 0), (-1, 0), 1, colors.white),
        ]

        if is_single_customer:
            # Left: Ref
            style_cmds.append(("ALIGN", (0, 1), (0, -1), "LEFT"))
            # Center: Doc, No, InvDate
            style_cmds.append(("ALIGN", (1, 1), (3, -1), "CENTER"))
            # Right: Total
            style_cmds.append(("ALIGN", (4, 1), (4, -1), "RIGHT"))
            # Center: ArrDate
            style_cmds.append(("ALIGN", (5, 1), (5, -1), "CENTER"))
            # Right: Pmt, Bal
            style_cmds.append(("ALIGN", (6, 1), (7, -1), "RIGHT"))
            # Center: DueDate, Days
            style_cmds.append(("ALIGN", (8, 1), (9, -1), "CENTER"))
        else:
            # Left: Cust, Ref
            style_cmds.append(("ALIGN", (0, 1), (1, -1), "LEFT"))
            # Center: Doc, No, InvDate
            style_cmds.append(("ALIGN", (2, 1), (4, -1), "CENTER"))
            # Right: Total
            style_cmds.append(("ALIGN", (5, 1), (5, -1), "RIGHT"))
            # Center: ArrDate
            style_cmds.append(("ALIGN", (6, 1), (6, -1), "CENTER"))
            # Right: Pmt, Bal
            style_cmds.append(("ALIGN", (7, 1), (8, -1), "RIGHT"))
            # Center: DueDate, Days
            style_cmds.append(("ALIGN", (9, 1), (10, -1), "CENTER"))

        t.setStyle(TableStyle(style_cmds))
        story.append(t)

        # --- Totales Currency ---
        cur_total = cur_group.totals['total']
        cur_pays = cur_group.totals['paid']
        cur_bal = cur_group.totals['balance']
        
        # Totales deben alinearse con las columnas Total, Pmt, Balance
        # Single: Total=4, Pmt=6, Bal=7. (Hay hueco en 5 ArrDate)
        # Multi: Total=5, Pmt=7, Bal=8. (Hay hueco en 6 ArrDate)
        
        if is_single_customer:
            # Row length = 10
            # Indices: 0,1,2,3 (Empty), 4 (Total), 5 (Empty), 6 (Pmt), 7 (Bal), 8,9 (Empty)
            # Label "TOTALS:" en 3 (InvDate column? No, mejor en 0-3 merged)
            totals_row_vals = [""] * 10
            totals_row_vals[3] = "TOTALS:" # En la columna previa a Total
            totals_row_vals[4] = f"{cur_total:,.2f}"
            totals_row_vals[6] = f"{cur_pays:,.2f}"
            totals_row_vals[7] = f"{cur_bal:,.2f}"
            
            span_range = (0, 3)
            label_idx = 3
            # Backgrounds en 4, 6, 7? O todo el bloque?
            # Excel suele poner background en los totales.
            # Vamos a poner background en 4, 6, 7
            bg_indices = [4, 6, 7]
        else:
            # Row length = 11
            # Indices: 0,1,2,3,4 (Empty), 5 (Total), 6 (Empty), 7 (Pmt), 8 (Bal), 9,10 (Empty)
            totals_row_vals = [""] * 11
            totals_row_vals[4] = "TOTALS:"
            totals_row_vals[5] = f"{cur_total:,.2f}"
            totals_row_vals[7] = f"{cur_pays:,.2f}"
            totals_row_vals[8] = f"{cur_bal:,.2f}"
            
            span_range = (0, 4)
            label_idx = 4
            bg_indices = [5, 7, 8]

        tt = Table([totals_row_vals], colWidths=[w * mm for w in widths])
        
        tt_cmds = [
            ("SPAN", (span_range[0], 0), (span_range[1], 0)),
            ("ALIGN", (label_idx, 0), (label_idx, 0), "RIGHT"),
            ("FONT", (label_idx, 0), (label_idx, 0), "Helvetica-Bold", 6),
        ]
        for idx_bg in bg_indices:
            tt_cmds.append(("BACKGROUND", (idx_bg, 0), (idx_bg, 0), total_color))
            tt_cmds.append(("FONT", (idx_bg, 0), (idx_bg, 0), "Helvetica-Bold", 6))
            tt_cmds.append(("ALIGN", (idx_bg, 0), (idx_bg, 0), "RIGHT"))
            
        tt.setStyle(TableStyle(tt_cmds))
        story.append(tt)
        story.append(Spacer(0, 6 * mm))

        # --- Summary Table ---
        story.append(Paragraph(f"SUMMARY — {cur}", title_style))
        
        if is_single_customer:
            hdr_summary = ["TOTAL BALANCE", "NOT DUE", "OVERDUE", "0-21", "22-30", "31-45", "45+"]
            widths_summary = [40, 30, 30, 30, 30, 30, 30]
        else:
            hdr_summary = ["CUSTOMER", "TOTAL", "NOT DUE", "OVERDUE", "0-21", "22-30", "31-45", "45+"]
            widths_summary = [65, 30, 30, 30, 30, 30, 30, 30]

        tbl_summary = [hdr_summary]

        grand_agg = AgingSummary(total_balance=0.0)

        for cust, agg in sorted(cur_group.aging_summary.items()):
            row_s = []
            if not is_single_customer:
                cust_display = cust or ""
                if len(cust_display) > 30:
                    cust_display = cust_display[:30] + "\n" + cust_display[30:50]
                row_s.append(cust_display)
            
            vals = [
                agg.total_balance, agg.not_yet_due, agg.overdue,
                agg.bucket_0_21, agg.bucket_22_30, agg.bucket_31_45, agg.bucket_45_plus
            ]
            row_s.extend([f"{v:,.2f}" for v in vals])
            tbl_summary.append(row_s)

            grand_agg.total_balance += agg.total_balance
            grand_agg.not_yet_due += agg.not_yet_due
            grand_agg.overdue += agg.overdue
            grand_agg.bucket_0_21 += agg.bucket_0_21
            grand_agg.bucket_22_30 += agg.bucket_22_30
            grand_agg.bucket_31_45 += agg.bucket_31_45
            grand_agg.bucket_45_plus += agg.bucket_45_plus

        t_summary = Table(tbl_summary, colWidths=[w * mm for w in widths_summary], repeatRows=1, splitByRow=True)
        
        # Estilos Summary
        # Si single customer: todo son números, alinear derecha.
        # Si no: primera col es texto (izq), resto números (der).
        align_start = 0 if is_single_customer else 1
        
        summary_style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), head_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 6),
            ("FONT", (0, 1), (-1, -1), "Helvetica", 6),
            ("GRID", (0, 0), (-1, -1), 0.25, line_color),
            ("ALIGN", (align_start, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, alt_color]),
        ]
        
        if not is_single_customer:
            summary_style_cmds.append(("ALIGN", (0, 1), (align_start - 1, -1), "LEFT"))

        t_summary.setStyle(TableStyle(summary_style_cmds))
        story.append(t_summary)

        # --- Totales Summary ---
        total_vals = [
            grand_agg.total_balance, grand_agg.not_yet_due, grand_agg.overdue,
            grand_agg.bucket_0_21, grand_agg.bucket_22_30, grand_agg.bucket_31_45, grand_agg.bucket_45_plus
        ]
        
        if is_single_customer:
            # Solo valores
            total_summary_row = [[f"{v:,.2f}" for v in total_vals]]
            tsr_style = [
                ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
                ("BACKGROUND", (0, 0), (-1, -1), total_color),
                ("FONT", (0, 0), (-1, -1), "Helvetica-Bold", 6),
            ]
        else:
            # "TOTALS:" + valores
            total_summary_row = [["TOTALS:"] + [f"{v:,.2f}" for v in total_vals]]
            tsr_style = [
                ("ALIGN", (0, 0), (0, 0), "RIGHT"),
                ("BACKGROUND", (0, 0), (-1, -1), total_color),
                ("FONT", (0, 0), (-1, -1), "Helvetica-Bold", 6),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]

        tsr = Table(total_summary_row, colWidths=[w * mm for w in widths_summary])
        tsr.setStyle(TableStyle(tsr_style))
        story.append(tsr)

    doc.build(story, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)

    # Rebobina el buffer y devuélvelo
    buffer.seek(0)
    return buffer

# app/reports/report_builder.py
# ... (después de la función create_pdf_report)

# --- NUEVA FUNCIÓN DE HTML ---

def create_html_report(
    data: Dict[str, CurrencyGroup], 
    logo_path: str, # Aún no lo usamos
    filters: dict
) -> io.BytesIO:
    """
    Toma los datos procesados y construye un archivo HTML EN MEMORIA.
    Devuelve un objeto io.BytesIO.
    """
    as_of = filters['as_of']
    customer_name = filters['customer_name']
    all_entries = [entry for group in data.values() for entry in group.entries]

    # Esta es tu lógica de CSS original
    css = """
    <style>
      :root{--p:#2E86AB;--pL:#F0F9FF;--s:#A7C957;--ink:#1F2937;--muted:#6B7280;--line:#D1D5DB;--alt:#F8F9FA}
      body{font-family:"Segoe UI",system-ui,-apple-system,sans-serif;color:var(--ink);margin:0;padding:24px}
      .header{display:flex;align-items:center;justify-content:space-between;border-bottom:3px solid var(--p);padding-bottom:16px;margin-bottom:20px}
      .title{color:var(--p);font-size:24px;font-weight:700;margin:0}
      .sub{color:var(--muted);margin:4px 0 0 0}
      .logo{height:46px}
      h2{color:var(--p)}
      table{width:100%;border-collapse:collapse;font-size:12px;margin:12px 0 28px}
      thead th{background:var(--p);color:#fff;text-transform:uppercase;font-size:11px;padding:8px;border:1px solid var(--p)}
      tbody td{padding:8px;border-bottom:1px solid var(--line)}
      tbody tr:nth-child(even){background:var(--alt)}
      .num{font-variant-numeric:tabular-nums; text-align:right;}
      .center{text-align:center;}
      .tot td{background:var(--s);color:#fff;font-weight:700}
      .meta{background:var(--pL);padding:10px;border-radius:6px;margin:10px 0 18px;display:flex;gap:24px}
      .meta b{color:var(--p)}
    </style>
    """

    logo_img = "" # Logo no implementado

    parts = [
        f"""<!doctype html><html><head><meta charset="utf-8"><title>AR Aging</title>{css}</head><body>
        <div class="header">{logo_img}<div style="flex:1;text-align:center">
          <div class="title">ACCOUNTS RECEIVABLE AGING</div>
          <div class="sub">Detailed Accounts Receivable Analysis</div>
        </div><div style="width:80px"></div></div>"""
    ]

    meta = f"""<div class="meta"><div>As Of: <b>{fmt_date_ordinal(as_of)}</b></div>
                <div>Customer: <b>{customer_name}</b></div>
                <div>Total Records: <b>{len(all_entries)}</b></div></div>"""
    parts.append(meta)

    # Detectar si es un solo cliente
    is_single_customer = filters.get('customer_id') is not None

    for cur, cur_group in data.items():
        parts.append(f"<h2>Currency — {cur}</h2>")
        
        if is_single_customer:
            # Excel Order: REFERENCE, DOCUMENT, NO., INVOICE DATE, TOTAL AMOUNT, ARRIVAL DATE, PAYMENTS, BALANCE, DUE DATE, DAYS SINCE ARRIVAL
            thead = (
                "<thead><tr>"
                "<th>REFERENCE</th><th>DOCUMENT</th><th>NO.</th><th>INVOICE DATE</th>"
                "<th>TOTAL AMOUNT</th><th>ARRIVAL DATE</th><th>PAYMENTS</th><th>BALANCE</th>"
                "<th>DUE DATE</th><th>DAYS SINCE ARRIVAL</th>"
                "</tr></thead>"
            )
        else:
            # Excel Order: CUSTOMER, REFERENCE, DOCUMENT, NO., INVOICE DATE, TOTAL AMOUNT, ARRIVAL DATE, PAYMENTS, BALANCE, DUE DATE, DAYS SINCE ARRIVAL
            thead = (
                "<thead><tr>"
                "<th>CUSTOMER</th><th>REFERENCE</th><th>DOCUMENT</th><th>NO.</th><th>INVOICE DATE</th>"
                "<th>TOTAL AMOUNT</th><th>ARRIVAL DATE</th><th>PAYMENTS</th><th>BALANCE</th>"
                "<th>DUE DATE</th><th>DAYS SINCE ARRIVAL</th>"
                "</tr></thead>"
            )

        parts.append(f"<table>{thead}<tbody>")
        
        for entry in cur_group.entries:
            row_html = "<tr>"
            
            if not is_single_customer:
                row_html += f"<td>{entry.customer_name or ''}</td>"
            
            row_html += f"<td>{entry.reference or ''}</td>"
            row_html += f"<td class='center'>{entry.module or ''}</td>"
            row_html += f"<td class='center'>{entry.folio or ''}</td>"
            row_html += f"<td class='center'>{fmt_date(entry.invoice_date)}</td>"
            row_html += f"<td class='num'>{entry.total:,.2f}</td>"
            row_html += f"<td class='center'>{fmt_date(entry.arrival_date)}</td>"
            row_html += f"<td class='num'>{entry.paid:,.2f}</td>"
            row_html += f"<td class='num'>{entry.balance:,.2f}</td>"
            row_html += f"<td class='center'>{fmt_date(entry.due_date)}</td>"
            row_html += f"<td class='num' style='text-align:center'>{entry.days_since}</td>"
            
            row_html += "</tr>"
            parts.append(row_html)

        cur_total = cur_group.totals['total']
        cur_pays = cur_group.totals['paid']
        cur_bal = cur_group.totals['balance']
        
        # Totales Row
        # Single: Total=col 5 (index 4), Pmt=col 7 (6), Bal=col 8 (7).
        # Multi: Total=col 6 (index 5), Pmt=col 8 (7), Bal=col 9 (8).
        
        if is_single_customer:
            # Colspan 4 (0-3) para label "TOTALS:"
            parts.append(
                f"<tr class='tot'><td colspan='4' style='text-align:right'>TOTAL ({cur}):</td>"
                f"<td class='num'>{cur_total:,.2f}</td>"
                f"<td></td>" # Arrival Date placeholder
                f"<td class='num'>{cur_pays:,.2f}</td>"
                f"<td class='num'>{cur_bal:,.2f}</td>"
                f"<td></td><td></td></tr>" # Due Date, Days placeholders
            )
        else:
            # Colspan 5 (0-4) para label "TOTALS:"
            parts.append(
                f"<tr class='tot'><td colspan='5' style='text-align:right'>TOTAL ({cur}):</td>"
                f"<td class='num'>{cur_total:,.2f}</td>"
                f"<td></td>" # Arrival Date placeholder
                f"<td class='num'>{cur_pays:,.2f}</td>"
                f"<td class='num'>{cur_bal:,.2f}</td>"
                f"<td></td><td></td></tr>" # Due Date, Days placeholders
            )
            
        parts.append("</tbody></table>")
        
        # --- Summary Table HTML ---
        parts.append(f"<h2>Summary — {cur}</h2>")
        
        if is_single_customer:
            thead_sum = (
                "<thead><tr>"
                "<th>TOTAL BALANCE</th><th>NOT YET DUE</th><th>OVERDUE</th>"
                "<th>0-21</th><th>22-30</th><th>31-45</th><th>45+ DAYS</th>"
                "</tr></thead>"
            )
        else:
            thead_sum = (
                "<thead><tr>"
                "<th>CUSTOMER</th><th>TOTAL BALANCE</th><th>NOT YET DUE</th><th>OVERDUE</th>"
                "<th>0-21</th><th>22-30</th><th>31-45</th><th>45+ DAYS</th>"
                "</tr></thead>"
            )
            
        parts.append(f"<table>{thead_sum}<tbody>")

        # (Usamos un dict simple para los totales, ya que no necesitamos el objeto Pydantic)
        grand_agg = {"total_balance": 0.0, "not_yet_due": 0.0, "overdue": 0.0, "bucket_0_21": 0.0, "bucket_22_30": 0.0, "bucket_31_45": 0.0, "bucket_45_plus": 0.0}

        for k, agg in sorted(cur_group.aging_summary.items()):
            vals = [
                agg.total_balance, agg.not_yet_due, agg.overdue,
                agg.bucket_0_21, agg.bucket_22_30, agg.bucket_31_45, agg.bucket_45_plus
            ]
            row_html = "<tr>"
            if not is_single_customer:
                row_html += "<td>" + (k or "") + "</td>"
            
            row_html += "".join(f"<td class='num'>{v:,.2f}</td>" for v in vals) + "</tr>"
            parts.append(row_html)

            grand_agg["total_balance"] += agg.total_balance
            grand_agg["not_yet_due"] += agg.not_yet_due
            grand_agg["overdue"] += agg.overdue
            grand_agg["bucket_0_21"] += agg.bucket_0_21
            grand_agg["bucket_22_30"] += agg.bucket_22_30
            grand_agg["bucket_31_45"] += agg.bucket_31_45
            grand_agg["bucket_45_plus"] += agg.bucket_45_plus

        tot = [
            grand_agg["total_balance"], grand_agg["not_yet_due"], grand_agg["overdue"],
            grand_agg["bucket_0_21"], grand_agg["bucket_22_30"], grand_agg["bucket_31_45"], grand_agg["bucket_45_plus"]
        ]
        
        if is_single_customer:
            parts.append(
                "<tr class='tot'>"
                + "".join(f"<td class='num'>{v:,.2f}</td>" for v in tot)
                + "</tr>"
            )
        else:
            parts.append(
                "<tr class='tot'><td>TOTALS:</td>"
                + "".join(f"<td class='num'>{v:,.2f}</td>" for v in tot)
                + "</tr>"
            )
            
        parts.append("</tbody></table>")

    parts.append("</body></html>")

    # --- ¡CAMBIO CLAVE! Guardar en memoria ---
    html_content = "".join(parts)
    buffer = io.BytesIO(html_content.encode('utf-8'))
    buffer.seek(0)
    return buffer